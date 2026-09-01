"""Excel bulk-import helpers for categories and products.

Managers can add many categories or products at once by uploading an Excel
(.xlsx) workbook. The import is transactional: if any row fails to validate,
nothing is written and a per-row error report is returned.

Supported column layouts (row 1 = headers, values start on row 2):

Categories sheet:
    Name | Name (Amharic) | Type

Products sheet (imported into an "Uncategorized" placeholder; the manager then
assigns category, price, sizes and option groups afterwards):
    Name | Name (Amharic) | Description | Description (Amharic)
"""
from datetime import time as time_of_day

from django.db import transaction

from .models import Category, Product


# Mapping of accepted column names (case-insensitive) -> canonical key.
CATEGORY_HEADER_MAP = {
    "name": "name",
    "name (amharic)": "name_amharic",
    "type": "type",
    "description": "description",
    "display order": "display_order",
    "available from": "available_from",
    "available to": "available_to",
}

PRODUCT_HEADER_MAP = {
    "name": "name",
    "name (amharic)": "name_amharic",
    "description": "description",
    "description (amharic)": "description_amharic",
}


def _header_row(ws, header_map):
    """Return a dict mapping canonical header key -> column index (0-based),
    or raise ValueError listing the missing required columns."""
    headers = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        key = header_map.get(str(cell.value).strip().lower())
        if key:
            headers[key] = cell.column - 1
    return headers


def _rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(value, row_num, col_name):
    text = _clean_text(value)
    if text == "":
        return None
    try:
        return float(text.replace(",", ""))
    except (TypeError, ValueError):
        raise ValueError(f"Row {row_num}: '{col_name}' must be a number.")


def _parse_time(value, row_num, col_name):
    if value is None or _clean_text(value) == "":
        return None
    if isinstance(value, time_of_day):
        return value
    if hasattr(value, "hour"):
        return time_of_day(value.hour, value.minute)
    text = str(value).strip()
    try:
        return time_of_day.fromisoformat(text)
    except ValueError:
        try:
            hour, minute = text.split(":")
            return time_of_day(int(hour), int(minute))
        except ValueError:
            raise ValueError(
                f"Row {row_num}: '{col_name}' must be a time like 08:00."
            )


CATEGORY_TEMPLATE_COLUMNS = ["Name", "Name (Amharic)"]
PRODUCT_TEMPLATE_COLUMNS = [
    "Name",
    "Name (Amharic)",
    "Description",
    "Description (Amharic)",
]


def build_category_template():
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    ws.append(CATEGORY_TEMPLATE_COLUMNS)
    ws.append([None, None])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_product_template():
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(PRODUCT_TEMPLATE_COLUMNS)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_categories(workbook, restaurant=None):
    """Import categories from an xlsx workbook. Returns (created_count, [ids], report)."""
    if "Categories" not in workbook.sheetnames:
        raise ValueError("The workbook must contain a sheet named 'Categories'.")
    ws = workbook["Categories"]
    headers = _header_row(ws, CATEGORY_HEADER_MAP)
    if "name" not in headers:
        raise ValueError("Categories sheet is missing a required 'Name' column.")

    rows = _rows(ws)
    created = 0
    ids = []
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            if not any(r is not None and _clean_text(r) for r in row):
                continue
            try:
                name = _clean_text(row[headers["name"]])
                if not name:
                    raise ValueError(f"Row {idx}: 'Name' is required.")
                if Category.objects.filter(name=name).exists():
                    raise ValueError(f"Row {idx}: A category named '{name}' already exists.")

                name_amharic = _clean_text(row[headers["name_amharic"]]) if "name_amharic" in headers else ""
                raw_type = _clean_text(row[headers["type"]]) if "type" in headers else ""
                category_type = raw_type.lower() if raw_type else "food"
                if category_type not in ("food", "drink"):
                    raise ValueError(f"Row {idx}: 'Type' must be 'food' or 'drink'.")
                description = _clean_text(row[headers["description"]]) if "description" in headers else ""
                display_order = _parse_number(row[headers["display_order"]], idx, "Display Order") if "display_order" in headers else None
                available_from = _parse_time(row[headers["available_from"]], idx, "Available From") if "available_from" in headers else None
                available_to = _parse_time(row[headers["available_to"]], idx, "Available To") if "available_to" in headers else None

                if (available_from is None) != (available_to is None):
                    raise ValueError(f"Row {idx}: Set both 'Available From' and 'Available To', or leave both empty.")
                if available_from is not None and available_from == available_to:
                    raise ValueError(f"Row {idx}: 'Available From' and 'Available To' cannot be the same.")

                category_obj = Category.objects.create(
                    restaurant=restaurant,
                    type=category_type,
                    name=name,
                    name_amharic=name_amharic,
                    description=description,
                    display_order=int(display_order) if display_order is not None else 0,
                    available_from=available_from,
                    available_to=available_to,
                )
                ids.append(category_obj.id)
                created += 1
            except ValueError as exc:
                errors.append(str(exc))

        if errors:
            transaction.set_rollback(True)
            created = 0
            ids = []

    return created, ids, errors


def import_products(workbook, restaurant=None):
    """Import products from an xlsx workbook.

    Only Name / Name (Amharic) / Description / Description (Amharic) are read.
    Newly created products are placed in an auto-created "Uncategorized"
    placeholder category so the manager can later assign category, price,
    sizes and option groups. Returns (created_count, [ids], errors).
    """
    if "Products" not in workbook.sheetnames:
        raise ValueError("The workbook must contain a sheet named 'Products'.")
    ws = workbook["Products"]
    headers = _header_row(ws, PRODUCT_HEADER_MAP)
    if "name" not in headers:
        raise ValueError("Products sheet is missing a required 'Name' column.")

    uncategorized, _ = Category.objects.get_or_create(
        name="Uncategorized",
        defaults={
            "restaurant": restaurant,
            "name_amharic": "አልተመደበም",
            "description": "Products waiting to be assigned to a category.",
            "display_order": 0,
        },
    )

    rows = _rows(ws)
    created = 0
    ids = []
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            if not any(r is not None and _clean_text(r) for r in row):
                continue
            try:
                name = _clean_text(row[headers["name"]])
                if not name:
                    raise ValueError(f"Row {idx}: 'Name' is required.")

                name_amharic = _clean_text(row[headers["name_amharic"]]) if "name_amharic" in headers else ""
                description = _clean_text(row[headers["description"]]) if "description" in headers else ""
                description_amharic = _clean_text(row[headers["description_amharic"]]) if "description_amharic" in headers else ""

                if Product.objects.filter(category=uncategorized, name=name).exists():
                    raise ValueError(f"Row {idx}: A product named '{name}' already exists in 'Uncategorized'.")

                product = Product.objects.create(
                    category=uncategorized,
                    name=name,
                    name_amharic=name_amharic,
                    description=description,
                    description_amharic=description_amharic,
                    price=0,
                    has_sizes=False,
                    is_active=True,
                )
                ids.append(product.id)
                created += 1
            except ValueError as exc:
                errors.append(str(exc))

        if errors:
            transaction.set_rollback(True)
            created = 0
            ids = []

    return created, ids, errors
